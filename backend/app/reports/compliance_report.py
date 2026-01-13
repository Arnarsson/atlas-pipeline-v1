"""
Compliance Report Generator

Generates compliance reports in PDF and Excel formats.
Includes GDPR compliance status, PII detections, quality scores,
and audit trail summaries.

EU AI Act Compliance: Art. 12 (Documentation), Art. 30 (Audit Logs)
"""

import io
import logging
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


class ReportFormat(str, Enum):
    """Supported report output formats."""
    PDF = "pdf"
    EXCEL = "excel"


class PIISummary(BaseModel):
    """Summary of PII detections."""
    total_detections: int = 0
    by_type: Dict[str, int] = Field(default_factory=dict)
    high_risk_count: int = 0
    masked_count: int = 0
    avg_confidence: float = 0.0


class QualitySummary(BaseModel):
    """Summary of data quality metrics."""
    overall_score: float = 0.0
    completeness: float = 0.0
    uniqueness: float = 0.0
    validity: float = 0.0
    timeliness: float = 0.0
    accuracy: float = 0.0
    consistency: float = 0.0
    issues_found: int = 0
    records_processed: int = 0


class GDPRStatus(BaseModel):
    """GDPR compliance status."""
    compliant: bool = True
    pending_requests: int = 0
    completed_requests: int = 0
    access_requests: int = 0
    deletion_requests: int = 0
    rectification_requests: int = 0
    retention_policy_days: int = 365
    data_processing_basis: str = "consent"


class AuditSummary(BaseModel):
    """Summary of audit events."""
    total_events: int = 0
    by_type: Dict[str, int] = Field(default_factory=dict)
    last_24h_events: int = 0
    critical_events: int = 0


class ReportData(BaseModel):
    """Data for compliance report generation."""
    report_id: str = Field(default_factory=lambda: f"rpt_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}")
    generated_at: datetime = Field(default_factory=datetime.utcnow)
    organization: str = "Atlas Intelligence"
    dataset_name: str = "Unknown"
    date_range_start: Optional[datetime] = None
    date_range_end: Optional[datetime] = None

    # Summaries
    pii: PIISummary = Field(default_factory=PIISummary)
    quality: QualitySummary = Field(default_factory=QualitySummary)
    gdpr: GDPRStatus = Field(default_factory=GDPRStatus)
    audit: AuditSummary = Field(default_factory=AuditSummary)

    # EU AI Act compliance
    eu_ai_act_risk_level: str = "limited"  # minimal, limited, high, unacceptable
    eu_ai_act_articles: List[str] = Field(default_factory=lambda: ["Art. 10", "Art. 11", "Art. 12", "Art. 30"])

    # Custom sections
    custom_sections: Dict[str, Any] = Field(default_factory=dict)


class ComplianceReportGenerator:
    """
    Generates compliance reports in PDF and Excel formats.

    Features:
    - GDPR compliance status
    - PII detection summary
    - Data quality metrics
    - Audit trail summary
    - EU AI Act compliance status
    """

    def __init__(self):
        """Initialize the report generator."""
        self._pdf_available = self._check_pdf_library()
        self._excel_available = self._check_excel_library()

        logger.info(
            f"ComplianceReportGenerator initialized: "
            f"PDF={self._pdf_available}, Excel={self._excel_available}"
        )

    def _check_pdf_library(self) -> bool:
        """Check if PDF generation library is available."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate
            return True
        except ImportError:
            logger.warning("reportlab not available - PDF generation disabled")
            return False

    def _check_excel_library(self) -> bool:
        """Check if Excel generation library is available."""
        try:
            import openpyxl
            return True
        except ImportError:
            logger.warning("openpyxl not available - Excel generation disabled")
            return False

    def generate(
        self,
        data: ReportData,
        format: ReportFormat = ReportFormat.PDF,
    ) -> bytes:
        """
        Generate a compliance report.

        Args:
            data: Report data
            format: Output format (PDF or Excel)

        Returns:
            Report as bytes
        """
        if format == ReportFormat.PDF:
            return self._generate_pdf(data)
        elif format == ReportFormat.EXCEL:
            return self._generate_excel(data)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_pdf(self, data: ReportData) -> bytes:
        """Generate PDF compliance report."""
        if not self._pdf_available:
            raise RuntimeError("PDF generation not available - install reportlab")

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30,
        )
        story.append(Paragraph("Compliance Report", title_style))

        # Report metadata
        meta_style = ParagraphStyle(
            'Meta',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
        )
        story.append(Paragraph(f"Report ID: {data.report_id}", meta_style))
        story.append(Paragraph(f"Generated: {data.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')}", meta_style))
        story.append(Paragraph(f"Organization: {data.organization}", meta_style))
        story.append(Paragraph(f"Dataset: {data.dataset_name}", meta_style))
        story.append(Spacer(1, 20))

        # Section: Executive Summary
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading1'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=10,
        )
        story.append(Paragraph("Executive Summary", heading_style))

        summary_data = [
            ["Metric", "Value", "Status"],
            ["Overall Quality Score", f"{data.quality.overall_score:.1%}", self._status_text(data.quality.overall_score >= 0.9)],
            ["PII Detections", str(data.pii.total_detections), self._status_text(data.pii.high_risk_count == 0)],
            ["GDPR Compliance", "Compliant" if data.gdpr.compliant else "Non-Compliant", self._status_text(data.gdpr.compliant)],
            ["EU AI Act Risk Level", data.eu_ai_act_risk_level.title(), self._status_text(data.eu_ai_act_risk_level in ["minimal", "limited"])],
        ]

        summary_table = Table(summary_data, colWidths=[200, 150, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))

        # Section: Data Quality
        story.append(Paragraph("Data Quality Metrics", heading_style))

        quality_data = [
            ["Dimension", "Score", "Threshold", "Status"],
            ["Completeness", f"{data.quality.completeness:.1%}", "95%", self._status_text(data.quality.completeness >= 0.95)],
            ["Uniqueness", f"{data.quality.uniqueness:.1%}", "98%", self._status_text(data.quality.uniqueness >= 0.98)],
            ["Validity", f"{data.quality.validity:.1%}", "90%", self._status_text(data.quality.validity >= 0.90)],
            ["Timeliness", f"{data.quality.timeliness:.1%}", "90%", self._status_text(data.quality.timeliness >= 0.90)],
            ["Accuracy", f"{data.quality.accuracy:.1%}", "90%", self._status_text(data.quality.accuracy >= 0.90)],
            ["Consistency", f"{data.quality.consistency:.1%}", "90%", self._status_text(data.quality.consistency >= 0.90)],
        ]

        quality_table = Table(quality_data, colWidths=[120, 100, 100, 130])
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(quality_table)
        story.append(Spacer(1, 30))

        # Section: PII Detection
        story.append(Paragraph("PII Detection Summary", heading_style))

        body_style = styles['Normal']
        story.append(Paragraph(f"Total PII Detections: {data.pii.total_detections}", body_style))
        story.append(Paragraph(f"High-Risk Findings: {data.pii.high_risk_count}", body_style))
        story.append(Paragraph(f"Masked Fields: {data.pii.masked_count}", body_style))
        story.append(Paragraph(f"Average Confidence: {data.pii.avg_confidence:.1%}", body_style))

        if data.pii.by_type:
            story.append(Spacer(1, 10))
            pii_type_data = [["PII Type", "Count"]]
            for pii_type, count in sorted(data.pii.by_type.items(), key=lambda x: -x[1]):
                pii_type_data.append([pii_type, str(count)])

            pii_table = Table(pii_type_data, colWidths=[200, 100])
            pii_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(pii_table)

        story.append(Spacer(1, 30))

        # Section: GDPR Status
        story.append(Paragraph("GDPR Compliance", heading_style))

        gdpr_data = [
            ["Metric", "Value"],
            ["Compliance Status", "Compliant" if data.gdpr.compliant else "Non-Compliant"],
            ["Data Processing Basis", data.gdpr.data_processing_basis.title()],
            ["Retention Policy", f"{data.gdpr.retention_policy_days} days"],
            ["Pending Requests", str(data.gdpr.pending_requests)],
            ["Completed Requests", str(data.gdpr.completed_requests)],
            ["Access Requests (Art. 15)", str(data.gdpr.access_requests)],
            ["Deletion Requests (Art. 17)", str(data.gdpr.deletion_requests)],
            ["Rectification Requests (Art. 16)", str(data.gdpr.rectification_requests)],
        ]

        gdpr_table = Table(gdpr_data, colWidths=[200, 200])
        gdpr_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(gdpr_table)
        story.append(Spacer(1, 30))

        # Section: EU AI Act
        story.append(Paragraph("EU AI Act Compliance", heading_style))

        story.append(Paragraph(f"Risk Classification: {data.eu_ai_act_risk_level.title()}", body_style))
        story.append(Paragraph("Applicable Articles:", body_style))
        for article in data.eu_ai_act_articles:
            story.append(Paragraph(f"  - {article}", body_style))

        story.append(Spacer(1, 30))

        # Section: Audit Trail
        story.append(Paragraph("Audit Trail Summary", heading_style))

        story.append(Paragraph(f"Total Audit Events: {data.audit.total_events}", body_style))
        story.append(Paragraph(f"Events (Last 24h): {data.audit.last_24h_events}", body_style))
        story.append(Paragraph(f"Critical Events: {data.audit.critical_events}", body_style))

        if data.audit.by_type:
            story.append(Spacer(1, 10))
            audit_type_data = [["Event Type", "Count"]]
            for event_type, count in sorted(data.audit.by_type.items(), key=lambda x: -x[1]):
                audit_type_data.append([event_type, str(count)])

            audit_table = Table(audit_type_data, colWidths=[200, 100])
            audit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(audit_table)

        # Footer
        story.append(Spacer(1, 50))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1,  # Center
        )
        story.append(Paragraph(
            f"Generated by Atlas Intelligence | {data.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')} | {data.report_id}",
            footer_style
        ))

        # Build PDF
        doc.build(story)

        buffer.seek(0)
        return buffer.getvalue()

    def _generate_excel(self, data: ReportData) -> bytes:
        """Generate Excel compliance report."""
        if not self._excel_available:
            raise RuntimeError("Excel generation not available - install openpyxl")

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='e2e8f0'),
            right=Side(style='thin', color='e2e8f0'),
            top=Side(style='thin', color='e2e8f0'),
            bottom=Side(style='thin', color='e2e8f0')
        )

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Compliance Report - Executive Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Metadata
        ws['A3'] = "Report ID:"
        ws['B3'] = data.report_id
        ws['A4'] = "Generated:"
        ws['B4'] = data.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')
        ws['A5'] = "Organization:"
        ws['B5'] = data.organization
        ws['A6'] = "Dataset:"
        ws['B6'] = data.dataset_name

        # Summary table
        summary_start = 8
        headers = ["Metric", "Value", "Threshold", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=summary_start, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        summary_rows = [
            ("Overall Quality Score", f"{data.quality.overall_score:.1%}", "90%", "PASS" if data.quality.overall_score >= 0.9 else "FAIL"),
            ("PII Detections", str(data.pii.total_detections), "N/A", "PASS" if data.pii.high_risk_count == 0 else "REVIEW"),
            ("High-Risk PII", str(data.pii.high_risk_count), "0", "PASS" if data.pii.high_risk_count == 0 else "FAIL"),
            ("GDPR Compliance", "Yes" if data.gdpr.compliant else "No", "Yes", "PASS" if data.gdpr.compliant else "FAIL"),
            ("EU AI Act Risk", data.eu_ai_act_risk_level.title(), "Limited", "PASS" if data.eu_ai_act_risk_level in ["minimal", "limited"] else "REVIEW"),
        ]

        for row_idx, row_data in enumerate(summary_rows, summary_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # Color status column
                if col_idx == 4:
                    if value == "PASS":
                        cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
                        cell.font = Font(color="166534")
                    elif value == "FAIL":
                        cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                        cell.font = Font(color="991b1b")
                    else:
                        cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
                        cell.font = Font(color="92400e")

        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

        # Sheet 2: Quality
        ws_quality = wb.create_sheet("Quality Metrics")

        ws_quality['A1'] = "Data Quality Metrics"
        ws_quality['A1'].font = Font(bold=True, size=14)

        quality_headers = ["Dimension", "Score", "Threshold", "Status"]
        for col, header in enumerate(quality_headers, 1):
            cell = ws_quality.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        quality_rows = [
            ("Completeness", data.quality.completeness, 0.95),
            ("Uniqueness", data.quality.uniqueness, 0.98),
            ("Validity", data.quality.validity, 0.90),
            ("Timeliness", data.quality.timeliness, 0.90),
            ("Accuracy", data.quality.accuracy, 0.90),
            ("Consistency", data.quality.consistency, 0.90),
        ]

        for row_idx, (name, score, threshold) in enumerate(quality_rows, 4):
            status = "PASS" if score >= threshold else "FAIL"
            row_data = [name, f"{score:.1%}", f"{threshold:.0%}", status]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_quality.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if col_idx == 4:
                    if value == "PASS":
                        cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
                        cell.font = Font(color="166534")
                    else:
                        cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                        cell.font = Font(color="991b1b")

        for col in range(1, 5):
            ws_quality.column_dimensions[get_column_letter(col)].width = 20

        # Sheet 3: PII
        ws_pii = wb.create_sheet("PII Detection")

        ws_pii['A1'] = "PII Detection Summary"
        ws_pii['A1'].font = Font(bold=True, size=14)

        ws_pii['A3'] = "Total Detections:"
        ws_pii['B3'] = data.pii.total_detections
        ws_pii['A4'] = "High-Risk Findings:"
        ws_pii['B4'] = data.pii.high_risk_count
        ws_pii['A5'] = "Masked Fields:"
        ws_pii['B5'] = data.pii.masked_count
        ws_pii['A6'] = "Avg Confidence:"
        ws_pii['B6'] = f"{data.pii.avg_confidence:.1%}"

        if data.pii.by_type:
            pii_headers = ["PII Type", "Count"]
            for col, header in enumerate(pii_headers, 1):
                cell = ws_pii.cell(row=8, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, (pii_type, count) in enumerate(sorted(data.pii.by_type.items(), key=lambda x: -x[1]), 9):
                ws_pii.cell(row=row_idx, column=1, value=pii_type).border = thin_border
                ws_pii.cell(row=row_idx, column=2, value=count).border = thin_border

        ws_pii.column_dimensions['A'].width = 25
        ws_pii.column_dimensions['B'].width = 15

        # Sheet 4: GDPR
        ws_gdpr = wb.create_sheet("GDPR Compliance")

        ws_gdpr['A1'] = "GDPR Compliance Status"
        ws_gdpr['A1'].font = Font(bold=True, size=14)

        gdpr_rows = [
            ("Compliance Status", "Compliant" if data.gdpr.compliant else "Non-Compliant"),
            ("Data Processing Basis", data.gdpr.data_processing_basis.title()),
            ("Retention Policy (Days)", data.gdpr.retention_policy_days),
            ("Pending Requests", data.gdpr.pending_requests),
            ("Completed Requests", data.gdpr.completed_requests),
            ("Access Requests (Art. 15)", data.gdpr.access_requests),
            ("Deletion Requests (Art. 17)", data.gdpr.deletion_requests),
            ("Rectification Requests (Art. 16)", data.gdpr.rectification_requests),
        ]

        for row_idx, (label, value) in enumerate(gdpr_rows, 3):
            ws_gdpr.cell(row=row_idx, column=1, value=label)
            ws_gdpr.cell(row=row_idx, column=2, value=value)

        ws_gdpr.column_dimensions['A'].width = 30
        ws_gdpr.column_dimensions['B'].width = 20

        # Sheet 5: Audit
        ws_audit = wb.create_sheet("Audit Trail")

        ws_audit['A1'] = "Audit Trail Summary"
        ws_audit['A1'].font = Font(bold=True, size=14)

        ws_audit['A3'] = "Total Events:"
        ws_audit['B3'] = data.audit.total_events
        ws_audit['A4'] = "Events (Last 24h):"
        ws_audit['B4'] = data.audit.last_24h_events
        ws_audit['A5'] = "Critical Events:"
        ws_audit['B5'] = data.audit.critical_events

        if data.audit.by_type:
            audit_headers = ["Event Type", "Count"]
            for col, header in enumerate(audit_headers, 1):
                cell = ws_audit.cell(row=7, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, (event_type, count) in enumerate(sorted(data.audit.by_type.items(), key=lambda x: -x[1]), 8):
                ws_audit.cell(row=row_idx, column=1, value=event_type).border = thin_border
                ws_audit.cell(row=row_idx, column=2, value=count).border = thin_border

        ws_audit.column_dimensions['A'].width = 25
        ws_audit.column_dimensions['B'].width = 15

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _status_text(self, is_good: bool) -> str:
        """Return status text."""
        return "PASS" if is_good else "REVIEW"


# Singleton instance
_generator: ComplianceReportGenerator | None = None


def get_report_generator() -> ComplianceReportGenerator:
    """Get or create the report generator instance."""
    global _generator
    if _generator is None:
        _generator = ComplianceReportGenerator()
    return _generator
